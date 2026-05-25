import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Load data and setup time horizons
lahore = pd.read_excel("dataset\pakistan_aqi_weather.xlsx", sheet_name="Lahore")
lahore['time'] = pd.to_datetime(lahore['time'])

# Establish categorical timeline horizons
lahore['Date'] = lahore['time'].dt.date
lahore['Month'] = lahore['time'].dt.to_period('M').astype(str)
lahore['Quarter'] = lahore['time'].dt.to_period('Q').astype(str)
lahore['Year'] = lahore['time'].dt.year

# Features to look across
features = ['pm2_5', 'temperature_2m', 'relative_humidity_2m', 'pressure_msl', 'wind_speed_10m', 'rain', 'cloud_cover']

def compute_profile_sheet(df, period_column, highest_label, lowest_label):
    """
    Computes a multi-index block profile matrix for all features across a given time horizon.
    """
    compiled_blocks = []
    
    for feature in features:
        # Calculate the contextual mean for this specific timeline horizon block
        horizon_means = df.groupby(period_column)[feature].transform('mean')
        
        # Segment observations based on the contextual mean threshold
        df['is_high'] = df[feature] >= horizon_means
        
        # Calculate components
        highest = df.groupby(period_column)[feature].max()
        lowest = df.groupby(period_column)[feature].min()
        overall_mean = df.groupby(period_column)[feature].mean()
        
        high_avg = df[df['is_high'] == True].groupby(period_column)[feature].mean()
        low_avg = df[df['is_high'] == False].groupby(period_column)[feature].mean()
        
        # Combine into a structured data segment block
        feature_df = pd.DataFrame({
            highest_label: highest,
            'high average': high_avg,
            'mean': overall_mean,
            'low average': low_avg,
            lowest_label: lowest
        }).T  # Transpose to pivot dates/periods to headers
        
        # Build multi-index rows (Feature Name -> Metric Profile Label)
        feature_df.index = pd.MultiIndex.from_product([[feature], feature_df.index], names=['Feature', 'Metric'])
        compiled_blocks.append(feature_df)
        
    # Concatenate all blocks vertically and sort columns chronologically
    final_profile = pd.concat(compiled_blocks)
    final_profile = final_profile.reindex(columns=sorted(final_profile.columns))
    final_profile.columns = final_profile.columns.astype(str)
    return final_profile

# 2. Compute the multi-metric profiles for all 4 horizons
print("Calculating profiles...")
daily_profile = compute_profile_sheet(lahore, 'Date', 'daily highest', 'daily lowest')
monthly_profile = compute_profile_sheet(lahore, 'Month', 'monthly highest', 'monthly lowest')
quarterly_profile = compute_profile_sheet(lahore, 'Quarter', 'quarterly highest', 'quarterly lowest')
yearly_profile = compute_profile_sheet(lahore, 'Year', 'yearly highest', 'yearly lowest')

# 3. Write raw profiles out to multi-tab excel sheet
excel_filename = "lahore_aggregation_report.xlsx"
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    daily_profile.to_excel(writer, sheet_name="Daily Profile")
    monthly_profile.to_excel(writer, sheet_name="Monthly Profile")
    quarterly_profile.to_excel(writer, sheet_name="Quarterly Profile")
    yearly_profile.to_excel(writer, sheet_name="Yearly Profile")

# 4. Apply professional presentation styling via OpenPyXL
print("Applying presentation styles...")
wb = openpyxl.load_workbook(excel_filename)

header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Navy Slate
header_font = Font(name="Segoe UI", size=11, bold=True, color="D9D9D9")
feature_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Muted Gray
feature_font = Font(name="Segoe UI", size=10, bold=True, color="D9D9D9")
metric_font = Font(name="Segoe UI", size=10, italic=False, color="D9D9D9")
data_font = Font(name="Segoe UI", size=10)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
)
thick_block_border = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='medium', color='1F4E79')
)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.views.sheetView[0].showGridLines = True  # Ensure gridlines remain visible
    
    # Format top headers (Dates/Horizons)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Format data rows and side multi-indexes
    for row_idx in range(2, ws.max_row + 1):
        is_block_end = (row_idx - 1) % 5 == 0
        current_border = thick_block_border if is_block_end else thin_border
        
        # Style Column A: Feature Groupings
        cell_a = ws.cell(row=row_idx, column=1)
        cell_a.font = feature_font
        cell_a.fill = feature_fill
        cell_a.alignment = Alignment(horizontal="left", vertical="center")
        cell_a.border = current_border
        
        # Style Column B: Metric Labels
        cell_b = ws.cell(row=row_idx, column=2)
        cell_b.font = metric_font
        cell_b.alignment = Alignment(horizontal="left", vertical="center")
        cell_b.border = current_border
        
        # Style Data Value Matrix (Column C onwards)
        for col_idx in range(3, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = current_border
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
    # Optimize column structural widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    for col_idx in range(3, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 13

wb.save(excel_filename)
print("Complete structured report generated successfully!")